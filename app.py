from flask import Flask, render_template, Response, request, jsonify
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from selenium import webdriver
import time
import glob
import shutil
import platform
import pymysql
import openpyxl
import pandas as pd
import numpy as np
import os
from openpyxl.styles import Color, PatternFill, Font, Border
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys


app = Flask(__name__)

COLOR_INDEX = (
    '00000000', '00FFFFFF', '00FF0000', '0000FF00', '000000FF', #0-4
    '00FFFF00', '00FF00FF', '0000FFFF', '00000000', '00FFFFFF', #5-9
    '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF', #10-14
    '0000FFFF', '00800000', '00008000', '00000080', '00808000', #15-19
    '00800080', '00008080', '00C0C0C0', '00808080', '009999FF', #20-24
    '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080', #25-29
    '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00', #30-34
    '0000FFFF', '00800080', '00800000', '00008080', '000000FF', #35-39
    '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF', #40-44
    '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC', #45-49
    '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699', #50-54
    '00969696', '00003366', '00339966', '00003300', '00333300', #55-59
    '00993300', '00993366', '00333399', '00333333',  #60-63
)


BLACK = COLOR_INDEX[0]
WHITE = COLOR_INDEX[1]
RED = COLOR_INDEX[2]
DARKRED = COLOR_INDEX[8]
BLUE = COLOR_INDEX[4]
DARKBLUE = COLOR_INDEX[12]
GREEN = COLOR_INDEX[3]
DARKGREEN = COLOR_INDEX[9]
YELLOW = COLOR_INDEX[5]
DARKYELLOW = COLOR_INDEX[19]

# conn = pymysql.connect('13.209.19.164', user='root', password='excel', db='coupang')
conn = pymysql.connect('localhost', user='root', password='P@ssw0rd', db='coupang')
winner_path = './files/winner.xlsx'
easy_path = './files/easy.xls'
inventory_path = "./files/price_inventory_200831_(1).xlsx"

output_dir = "./output/"

from datetime import datetime

#SELECT LOG_TYPE, COUNT(LOG_TYPE) AS count FROM `log` WHERE `LOG_TIME` BETWEEN '2020-08-27 00:00:00' AND '2020-08-27 23:59:59' GROUP BY LOG_TYPE

def log(log_type):
    curs = conn.cursor()
    now = datetime.now()  # current date and time

    date_time = now.strftime("%Y-%m-%d %H:%M:%S")

    curs.execute("INSERT INTO log(LOG_TYPE, LOG_TIME) VALUES (%s, %s)", (log_type, date_time))
    conn.commit()


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/db')
def db():
    return render_template('db.html')

@app.route('/getRel', methods=['POST'])
def getRel():
    curs = conn.cursor(pymysql.cursors.DictCursor)
    query = "SELECT * FROM rel"
    curs.execute(query)
    result = curs.fetchall()
    return jsonify(result)


@app.route('/deleteRel', methods=['POST'])
def deleteRel():
    curs = conn.cursor(pymysql.cursors.DictCursor)
    opt_ids = request.form.getlist("opt_id[]")
    query = "DELETE FROM rel WHERE OPT_ID=%s"
    for opt_id in opt_ids:

        curs.execute(query, opt_id)
    conn.commit()
    return "Success"


@app.route('/updateRel', methods=['POST'])
def updateRel():
    curs = conn.cursor(pymysql.cursors.DictCursor)
    opt_id = request.form.get("opt_id")
    col = request.form.get("col")
    data = request.form.get("data")

    update_query = "UPDATE rel SET {}=%s WHERE OPT_ID=%s".format(col)
    curs.execute(update_query, (data, opt_id))
    conn.commit()
    return "Success"

@app.route('/getLog')
def getLog():
    curs = conn.cursor(pymysql.cursors.DictCursor)

    result = {}

    today = request.args.get('date')
    count_query = "SELECT LOG_TYPE, COUNT(LOG_TYPE) AS COUNT FROM `log` WHERE `LOG_TIME` BETWEEN %s AND %s GROUP BY LOG_TYPE"
    curs.execute(count_query, ("{} 00:00:00".format(today), "{} 23:59:59".format(today)))
    counts = curs.fetchall()
    result['counts'] = counts

    select_query = "SELECT * FROM `log` WHERE `LOG_TIME` BETWEEN %s AND %s ORDER BY LOG_TIME ASC"
    curs.execute(select_query, ("{} 00:00:00".format(today), "{} 23:59:59".format(today)))
    logs = curs.fetchall()
    result['logs'] = logs

    return jsonify(result)


@app.route('/checkFilesWinner')
def checkFilesWinner():
    filelist = glob.glob(".\\files/*.*")
    if winner_path.replace('/', '\\') not in filelist:
        return "winner"
    if easy_path.replace('/', '\\') not in filelist:
        return "easy"
    if inventory_path.replace('/', '\\') not in filelist:
        return "inventory"
    return "success"



@app.route('/checkFilesStock')
def checkFilesStock():
    filelist = glob.glob(".\\files/*.*")
    if easy_path.replace('/', '\\') not in filelist:
        return "easy"
    if inventory_path.replace('/', '\\') not in filelist:
        return "inventory"
    return "success"



@app.route('/cafe24Option')
def cafe24Option():
    easy_path = './files/easy2.xlsx'
    df = pd.read_excel(easy_path)
    
    df1 = df[['상품코드','상품명','정상재고']]
    df2 = df[['옵션추가항목1','상품코드','상품명','정상재고']]
    df2.columns = ['상품코드','상품코드2','상품명2','정상재고2']

    df_Full = pd.merge(df1,df2, on = '상품코드')
    df_Full['작업'] = "M"

    df_finish = df_Full[['상품코드2','상품명2','정상재고','작업']]
    
    def change(x):
        if x == 0 : return "zero"
        else: return x

    df_finish["정상재고"] = df_finish["정상재고"].apply(change)
        

    df_finish.columns = ['상품코드','상품명','작업수량','작업']
    
    df_final = df_finish[['상품코드','작업수량','작업']]

    df_final.to_excel('./ea_upload.xlsx', index=False)
   
    return "ok"



   
@app.route('/cafe24Stock')
def cafe24Stock():
    options = Options()
    options.add_argument('--start-fullscreen')
    # options.add_argument('headless')
    chromedriver_path = './chromedriver'
    # driver = webdriver.Chrome(chromedriver_path, chrome_options=options)
    link = "https://www.ezadmin.co.kr/index.html"
    driver = webdriver.Chrome(chromedriver_path, chrome_options=options)
    driver.get(link)
    driver.find_element_by_xpath('//*[@id="header"]/div[2]/ul/li[2]').click()

    domain = 'beautystreet2'
    id = 'sure1782'
    pwd = 'zmffldh070@'
    
    driver.execute_script(f"document.getElementsByName('domain')[0].value='{domain}'")
    driver.execute_script(f"document.getElementsByName('userid')[0].value='{id}'")
    driver.execute_script(f"document.getElementsByName('passwd')[0].value='{pwd}'")



    page2 = driver.page_source
    bs = BeautifulSoup(page2,  "html.parser")
    driver.find_element_by_xpath('//*[@id="login-popup"]/div[2]/form[2]/input[4]').click()
    driver.find_element_by_xpath('//*[@id="passwd_keep"]/a[3]/img').click()
    driver.find_element_by_xpath('//*[@id="pop_top"]/span/a/img').click()
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="gnb"]/h2[6]/a').click()
    time.sleep(3)
    driver.find_element_by_xpath('//*[@id="sideMenu"]/dl/dd[10]/a/div').click()
    time.sleep(1)
    
    driver.find_element_by_xpath('//*[@id="stock_type"]/option[4]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="except_soldout"]/option[1]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="search"]').click()
    time.sleep(1)

    
    driver.find_element_by_xpath('//*[@id="wrap_grid_button"]/a[1]/span').click()
    time.sleep(2)

    page2 = driver.page_source
    bs = BeautifulSoup(page2,  "html.parser")
    how = bs.select_one('#zb_waiting')
    print(how)
    Hi = 'hidden'
    
    while True:
        page2 = driver.page_source
        bs = BeautifulSoup(page2,  "html.parser")
        how = bs.select_one('#zb_waiting')
        print(how)
        time.sleep(2)
        if Hi in str(how):
            break

    driver.close()


    return "ok"

@app.route('/cafe24Modify')
def cafe24Modify():
    def makecafe24Modify():
        options = Options()
        options = webdriver.ChromeOptions()
        # options.add_argument('headless')
        chromedriver_path = './chromedriver'
        driver = webdriver.Chrome(chromedriver_path, chrome_options=options)
        # label.config(text="btn2, Clicked!")
        href = "https://eclogin.cafe24.com/Shop/"     
        # chromedriver_path = 'C:/python-program/webcrawling_1230-r2/chromedriver'
        # driver = webdriver.Chrome(chromedriver_path)
        driver.get(href)
        time.sleep(2)
    
        # 카페24 아이디 / 비번 입력
        driver.find_element_by_id('mall_id').send_keys('kkwjkd')
        driver.find_element_by_id('userpasswd').send_keys('zmffldh123')
        driver.find_element_by_class_name('btnSubmit').click()
        time.sleep(7)

        yield "data:10\n\n"
        #팝업삭제
        try:
            driver.find_element_by_xpath('//*[@id="admngLayerWrapper29"]/form/div/button').click()
            time.sleep(2)
        except:
            pass
    

        #상품관리 버튼클릭
        driver.find_element_by_xpath('//*[@id="QA_Gnb_product2"]').click()
        time.sleep(2)

        #재고관리 버튼클릭
        driver.find_element_by_xpath('//*[@id="QA_Lnb_Menu2049"]').click()
        time.sleep(2)


        driver.find_element_by_xpath('//*[@id="QA_Lnb_Menu2050"]').click()
        time.sleep(2)

        #카테고리 화장품으로 수정
        driver.find_element_by_xpath('//*[@id="eCategory1"]/option[2]').click()
        time.sleep(2)

        #하위분류 포함검색 체크
        driver.find_element_by_xpath('//*[@id="submitSearchBox"]/table/tbody/tr[2]/td/div/span/label[1]/input').click()
        time.sleep(1)

        #상세검색열기 클릭
        driver.find_element_by_xpath('//*[@id="QA_list1"]/div[4]/div/span/button').click()
        time.sleep(1)


        #재고관리 사용안함 체크
        driver.find_element_by_xpath('//*[@id="QA_list1"]/div[3]/table/tbody/tr[1]/td/label[3]/input').click()
        time.sleep(1)
        
        #100개씩 보기선택
        driver.find_element_by_xpath('//*[@id="QA_list2"]/div[2]/div[2]/select[2]/option[5]').click()
        time.sleep(2)

        #검색버튼 클릭
        driver.find_element_by_xpath('//*[@id="eBtnSearch"]/span').click()
        time.sleep(2)
    
        yield "data:30\n\n"
        #사용안함을 사용함,품절사용체크 반복문
        i = 1   
        while True:
            try:

                
                # paging = driver.find_element_by_xpath('//*[@id="QA_list2"]/div[6]/ol/li['+str(i)+']')
                # print(paging.text)

                #체크박스 체크
                driver.find_element_by_xpath('//*[@id="QA_list2"]/div[4]/table/thead/tr/th[5]/input').click()
                
                #재고관리 일괄설정클릭
                driver.find_element_by_xpath('//*[@id="QA_list2"]/div[3]/div[1]/a[1]').click()

                #팝업창 열리고 사용함 선택
                driver.find_element_by_xpath('//*[@id="eManageStockBatchForm"]/table/tbody/tr[1]/td[1]/select/option[1]').click()
                
                #품절가능 체크
                driver.find_element_by_xpath('//*[@id="eManageStockBatchForm"]/table/tbody/tr[1]/td[6]/input').click()

                #선택완료버튼 클릭
                driver.find_element_by_xpath('//*[@id="layerBatchSet"]/div[2]/a[1]').click()
                time.sleep(2)
                #팝업창 완료
                driver.switch_to_alert().accept()
                time.sleep(2)

                #팝업창 문자확인
                # poptext = driver.switch_to_alert().text
                # if poptext == "처리할 품목이 없습니다.":
                #     break
                
                #팝업창 완료
                driver.switch_to_alert().accept()
                time.sleep(3)

                # driver.switch_to_window
                # time.sleep(3)
                continue
                # i += 1
                # if i == 11:
                #     driver.find_element_by_css_selector('#QA_list2 > div.mPaginate > a.next').click()
                #     i = 1
                #     continue
                    
            except:
                break
    



        # 사용함으로 모두 수정후 닫기 클릭
        driver.find_element_by_xpath('//*[@id="layerBatchSet"]/div[2]/a[2]/span').click()
        time.sleep(3)

        #상세검색창에 재고 사용함 체크
        driver.find_element_by_xpath('//*[@id="QA_list1"]/div[3]/table/tbody/tr[1]/td/label[2]/input').click()
        time.sleep(1)

        #재고 0부터
        driver.find_element_by_class_name('fText.right.eSearchText').send_keys('0')
        time.sleep(1)

        #재고 0까지입력
        driver.find_element_by_css_selector('#eSearchFormStock > li > input:nth-child(3)').send_keys('0')
        time.sleep(1)

        #판매상태 판매함 체크
        driver.find_element_by_xpath('//*[@id="submitSearchBox"]/table/tbody/tr[4]/td[2]/label[2]/input').click()
        time.sleep(1)

        #진열상태 진열함 체크
        driver.find_element_by_xpath('//*[@id="QA_list1"]/div[3]/table/tbody/tr[5]/td[1]/label[2]/input').click()
        time.sleep(1)


        #검색버튼 클릭
        driver.find_element_by_xpath('//*[@id="eBtnSearch"]').click()
        
        yield "data:60\n\n"
        #재고 0 상품을 재고관리 사용안함으로 수정 반복문
        while True:
            try:
                #체크버튼 전체선택
                driver.find_element_by_xpath('//*[@id="QA_list2"]/div[4]/table/thead/tr/th[5]/input').click()
                time.sleep(2)
                
                #재고관리 일괄설정
                driver.find_element_by_xpath('//*[@id="QA_list2"]/div[3]/div[1]/a[1]').click()
                time.sleep(2)

                #재고관리 사용안함 체크
                driver.find_element_by_xpath('//*[@id="eManageStockBatchForm"]/table/tbody/tr[1]/td[1]/select/option[2]').click()
                time.sleep(2)

                #진열여부체크
                driver.find_element_by_xpath('//*[@id="eManageStockBatchForm"]/table/tbody/tr[2]/td/div/table/tbody/tr[1]/th/label/input').click()
                time.sleep(1)

                #판매여부체크
                driver.find_element_by_xpath('//*[@id="eManageStockBatchForm"]/table/tbody/tr[2]/td/div/table/tbody/tr[2]/th/label/input').click()
                time.sleep(1)




                #확인 버튼 클릭
                driver.find_element_by_xpath('//*[@id="layerBatchSet"]/div[2]/a[1]/span').click()
                time.sleep(2)

                #팝업창 확인 클릭
                driver.switch_to_alert().accept()
                time.sleep(2)

                #팝업창 확인 클릭
                driver.switch_to_alert().accept()
                time.sleep(2)
                # driver.switch_to_window
                time.sleep(2)
                continue

            except:
                break

        

        # 사용안함으로 모두 수정후 닫기 클릭
        driver.find_element_by_xpath('//*[@id="layerBatchSet"]/div[2]/a[2]/span').click()
        time.sleep(3)


        try:
        #팝업창 제거
            driver.find_element_by_xpath('//*[@id="layerBatchSet"]/div[2]/a[2]').click()
            time.sleep(1)

        except:
            pass


        #재고0 상품 모두 품절 처리

        yield "data:80\n\n"
        #상품관리 메뉴클릭
        driver.find_element_by_xpath('//*[@id="QA_Lnb_Menu2036"]').click()
        time.sleep(1)

        #상품목록 메뉴클릭
        driver.find_element_by_xpath('//*[@id="QA_Lnb_Menu2037"]').click()
        time.sleep(1)

        
        #상세검색클릭
        driver.find_element_by_xpath('//*[@id="QA_list1"]/div[4]/div/span/button').click()
        time.sleep(2)

        #카테고리 화장품으로 수정
        driver.find_element_by_xpath('//*[@id="eCategory1"]/option[2]').click()
        time.sleep(2)

        #하위분류 포함검색 체크
        driver.find_element_by_xpath('//*[@id="submitSearchBox"]/table/tbody/tr[3]/td/div/span/label[1]/input').click()
        time.sleep(1)

        

        #재고관리 사용안함 체크
        driver.find_element_by_xpath('//*[@id="QA_list1"]/div[3]/table/tbody/tr[1]/td/label[3]/input').click()
        time.sleep(2)

        #판매상태 판매함으로 체크
        driver.find_element_by_xpath('//*[@id="submitSearchBox"]/table/tbody/tr[5]/td[2]/label[2]/input').click()
        time.sleep(2)

        #상품 100개 열기
        driver.find_element_by_xpath('//*[@id="QA_list2"]/div[2]/div[2]/select[2]/option[5]').click()
        time.sleep(2)

        #검색버튼 클릭
        driver.find_element_by_xpath('//*[@id="eBtnSearch"]/span').click()
        time.sleep(2)


        #재고관리 사용안함, 판매함으로 되어 있는 상품들 모두 판매안함으로 수정
        
        while True:
            try:

                #상품 체크박스 선택
                driver.find_element_by_xpath('//*[@id="QA_list2"]/div[4]/table/thead/tr/th[1]/input').click()
                time.sleep(2)

                #판매안함 버튼 클릭
                driver.find_element_by_xpath('//*[@id="QA_list2"]/div[3]/div[1]/a[4]/span').click()
                time.sleep(2)

                #팝업창 확인
                driver.switch_to_alert().accept()
                time.sleep(2)

                #팝업창 확인
                driver.switch_to_alert().accept()
                time.sleep(2)
                # driver.switch_to_window
                time.sleep(2)
                continue
            except:
                break

        time.sleep(3)
        try: 
            driver.switch_to_alert().accept()
            time.sleep(2)

        except:
            pass

        #입고상품 판매함으로 수정

        
        #재고관리 전체로 선택
        driver.find_element_by_xpath('//*[@id="QA_list1"]/div[3]/table/tbody/tr[1]/td/label[1]/input').click()
        time.sleep(2)

        #판매상태 판매안함 선택
        driver.find_element_by_xpath('//*[@id="submitSearchBox"]/table/tbody/tr[5]/td[2]/label[3]/input').click()
        time.sleep(2)

        #재고관리 재고 1입력 
        driver.find_element_by_css_selector('#eSearchFormStock > li > input:nth-child(2)').send_keys('1')
        time.sleep(2)

        #검색버튼 클릭
        driver.find_element_by_xpath('//*[@id="eBtnSearch"]/span').click()
        time.sleep(4)

        yield "data:90\n\n"
        
        #판매안함 상품을 모두 판매함으로 수정
        while True:
            try:

                #체크박스 선택
                driver.find_element_by_xpath('//*[@id="QA_list2"]/div[4]/table/thead/tr/th[1]/input').click()
                time.sleep(2)

                #판매함버튼 클릭
                driver.find_element_by_xpath('//*[@id="QA_list2"]/div[3]/div[1]/a[3]/span').click()
                time.sleep(2)

                #팝업창 확인
                driver.switch_to_alert().accept()
                time.sleep(2)
                
                #팝업창 확인
                driver.switch_to_alert().accept()
                time.sleep(2)
                driver.switch_to_window
                time.sleep(2)
                continue
            except:
                break

        
        driver.close()
        yield "data:100\n\n"
    return Response(makecafe24Modify(), mimetype='text/event-stream')
        




@app.route('/getStock')
def getStock():
    log(1)
    def makeStock():
        file_coupang = inventory_path

        data = pd.read_html(easy_path, header=0)
        df_easy = data[0]

        ##
        ## price_inventory에서 코드_수량을 DB에 추가
        ##
        whole = pd.read_excel(inventory_path, dtype={'옵션 ID': str, '업체상품코드': str}, header=2)
        add = whole[~whole["업체상품코드"].isnull() & whole["업체상품코드"].str.contains("_")]
        data = add.loc[:, ["옵션 ID", "업체상품코드", "쿠팡 노출 상품명"]]

        def parseCode(x):
            codes = str(x).split(",")
            codes = [code.strip() for code in codes]
            cs = []
            qs = 0
            for code in codes:
                if "_" in code:
                    c, q = code.split("_")
                    cs.append(c)
                    qs += int(q)
                else:
                    cs.append(code)
                    qs += 0
            return ",".join(cs), qs

        data["상품코드"], data["수량"] = zip(*data["업체상품코드"].map(parseCode))
        curs = conn.cursor()

        select_query = "SELECT * FROM rel WHERE OPT_ID=%s"
        row_data = data.loc[:, ["옵션 ID", "쿠팡 노출 상품명", "상품코드", "수량"]].values.tolist()
        insert_data = []
        for idx, opt_id in enumerate(data.loc[:, ["옵션 ID"]].values.tolist()):

            curs.execute(select_query, opt_id)
            result = curs.fetchall()
            if len(result) == 0:
                insert_data.append(row_data[idx])
        insert_query = "INSERT INTO rel VALUES (%s, %s, %s, %s)"
        for d in insert_data:
            curs.execute(insert_query, d)
        conn.commit()
        yield "data:20\n\n"
        ##
        ## price.xlsx 만드는 부분
        ##


        df_coupang = whole

        df_coupang["상품코드"], df_coupang["상품명수량"] = zip(*df_coupang["업체상품코드"].map(parseCode))


        curs = conn.cursor(pymysql.cursors.DictCursor)
        curs.execute("SELECT * FROM rel")
        result = curs.fetchall()
        df_connect = pd.DataFrame(result)

        def getEasy(opt_id):
            codes = df_connect[df_connect["OPT_ID"] == opt_id]["CODE"].values
            codes = list(set(list(codes)))
            codes = [code for code in codes if type(code) == str]
            if len(codes) == 0:
                return ""

            codes = codes[0]
            result = []
            for code in codes.split(","):
                result += df_easy[df_easy["상품코드"] == code]["정상재고"].values.tolist()
            # result : ["12", "6"]

            if 1 < len(result):
                result = min(result)
                return result
            else:
                result = [str(r) for r in result]
                return ",".join(result)
                #     return result

        def getQty(opt_id):
            codes = df_connect[df_connect["옵션ID"] == opt_id]["옵션ID"].values
            codes = list(set(list(codes)))
            codes = [code for code in codes if type(code) == str]
            if len(codes) == 0:
                return ""

            codes = codes[0]
            result = []
            for code in codes.split(","):
                result += df_connect[df_connect["옵션ID"] == code]["상품수량"].values.tolist()
            # result : ["12", "6"]
            result = [str(r) for r in result]
            return ",".join(result)

        def str2int(x):
            try:
                return int(x)
            except:
                return np.nan

        df_coupang[["실제재고"]] = df_coupang[["옵션 ID"]].applymap(getEasy)

        df_coupang["판매상태.1"] = df_coupang["잔여수량"].apply(lambda ea: "판매중지" if str(ea) == '0' else "판매중")

        df_conn = df_connect[["OPT_ID", 'QT']]
        df_conn.columns = ["옵션 ID", "상품수량"]

        df_coupang = pd.merge(df_coupang, df_conn, how='left', on='옵션 ID')

        df_coupang['실제재고'] = df_coupang['실제재고'].apply(str2int)
        df_coupang['상품수량'] = df_coupang['상품수량'].apply(str2int)

        df_coupang["최종재고"] = df_coupang["실제재고"].div(df_coupang["상품수량"])

        df_coupang["최종재고"] = df_coupang["최종재고"].round(0)

        df_coupang["잔여수량"] = df_coupang["최종재고"]
        df_coupang = df_coupang.drop(["상품코드", "상품명수량"], axis=1)
        df_coupang.to_excel(os.path.join(output_dir, 'ea_final.xlsx'), index=False)

        wb = openpyxl.load_workbook(os.path.join(output_dir, 'ea_final.xlsx'))
        ws = wb.active

        colors = {"판매상태.1": YELLOW, "잔여수량": GREEN,
                  "실제재고": YELLOW, "상품수량": YELLOW,
                  "최종재고": GREEN}

        for col in ws.columns:
            if col[0].value in colors:
                letter = col[0].column_letter
                for c in ws[letter]:
                    c.fill = openpyxl.styles.PatternFill('solid', colors[col[0].value])

        wb.save(filename=os.path.join(output_dir, 'ea_final.xlsx'))
        time.sleep(5)
        yield "data:99\n\n"
        time.sleep(0.5)
        yield "data:100\n\n"
    return Response(makeStock(), mimetype='text/event-stream')

@app.route('/getWinner')
def getWinner():
    log(0)

    def makeWinner():
        ##
        ## price_inventory에서 코드_수량을 DB에 추가
        ##
        add = pd.read_excel(inventory_path, dtype={'옵션 ID': str, '업체상품코드': str}, header=2)
        add = add[~add["업체상품코드"].isnull() & add["업체상품코드"].str.contains("_")]
        data = add.loc[:, ["옵션 ID", "업체상품코드", "쿠팡 노출 상품명"]]

        def parseCode(x):
            codes = x.split(",")
            codes = [code.strip() for code in codes]
            cs = []
            qs = 0
            for code in codes:
                c, q = code.split("_")
                cs.append(c)
                qs += int(q)
            return ",".join(cs), qs

        data["상품코드"], data["수량"] = zip(*data["업체상품코드"].map(parseCode))
        curs = conn.cursor()

        select_query = "SELECT * FROM rel WHERE OPT_ID=%s"
        row_data = data.loc[:, ["옵션 ID", "쿠팡 노출 상품명", "상품코드", "수량"]].values.tolist()
        insert_data = []
        for idx, opt_id in enumerate(data.loc[:, ["옵션 ID"]].values.tolist()):

            curs.execute(select_query, opt_id)
            result = curs.fetchall()
            if len(result) == 0:
                insert_data.append(row_data[idx])
        insert_query = "INSERT INTO rel VALUES (%s, %s, %s, %s)"
        for d in insert_data:
            curs.execute(insert_query, d)
        conn.commit()
        yield "data:20\n\n"
        ##
        ## price.xlsx 만드는 부분
        ##

        df_coupang = pd.read_excel(winner_path)
        df_coupang.loc[:, "옵션ID"] = df_coupang.loc[:, "옵션ID"].apply(lambda x: str(x).split(".")[0])

        data = pd.read_html(easy_path, header=0)
        df = data[0]
        df_easy = df.loc[:, ["상품코드", "원가"]]

        def Changecode(code):
            if len(code) == 5:
                code = str(code)
            elif len(code) == 4:
                code = str("0" + str(code))
            else:
                code = str("00" + str(code))

            return code

        df_easy["상품코드"] = df_easy["상품코드"].apply(Changecode)

        curs = conn.cursor(pymysql.cursors.DictCursor)
        curs.execute("SELECT * FROM rel")
        result = curs.fetchall()
        df_connect = pd.DataFrame(result)

        def getEasy(opt_id):
            codes = df_connect[df_connect["OPT_ID"] == opt_id]["CODE"].values
            codes = list(set(list(codes)))
            codes = [code for code in codes if type(code) == str]
            if len(codes) == 0:
                return ""

            codes = codes[0]
            result = []
            for code in codes.split(","):
                result += df_easy[df_easy["상품코드"] == code]["원가"].values.tolist()
            result = [str(r) for r in result]
            return ",".join(result)

        df_connect[["원가"]] = df_connect[["OPT_ID"]].applymap(getEasy)

        df_conn = df_connect[["OPT_ID", '원가', 'QT']]
        df_conn.columns = ["옵션ID", "원가", "상품수량"]
        yield "data:35\n\n"

        def setPrice(data):
            x = str(data)
            if x == '':
                return ''
            xs = x.split(",")
            result = 0
            for d in xs:

                if d == 'nan':
                    result += 0
                else:
                    result += int(d.split(".")[0])
            if result == 0:
                return ''
            return str(result)

        df_coupang = pd.merge(df_coupang, df_conn, how='left', on='옵션ID')
        yield "data:40\n\n"

        def float2int(x):
            if str(x) == 'nan':
                return 0
            else:
                return int(x)


        def float3int(x):
            try:
                return str(round(int(x), 2))
            except:
                return ""

        df_coupang["상품수량"] = df_coupang["상품수량"].apply(float2int)
        df_coupang[['원가', '상품수량', '판매자 배송비(원)']].dtypes

        df_coupang.loc[:, "원가"] = df_coupang.loc[:, "원가"].apply(setPrice)

        def str2int(x):
            try:
                return int(x)
            except:
                return np.nan

        df_coupang['원가'] = df_coupang['원가'].apply(str2int)

        low_price = []
        for value in df_coupang[['원가', '상품수량', '판매자 배송비(원)']].values:
            t = value[0] * value[1]
            if value[2] > 0:
                low_price.append(t / 0.89 + value[2])
            else:
                low_price.append((t + 1700) / 0.89)
        df_coupang["노마진"] = low_price
        yield "data:50\n\n"
        df_coupang['넷토피아가격'] = df_coupang['판매자 판매가격(원)'] + df_coupang['판매자 배송비(원)']
        df_coupang['쿠팡위너가격'] = df_coupang['아이템위너 판매가격(원)'] + df_coupang['아이템위너 배송비(원)']
        prices = []
        text = []
        for value in df_coupang[['넷토피아가격', '쿠팡위너가격', '판매자 배송비(원)', '노마진']].values:
            #   넷토피아 가격 <= 쿠팡위너가격 보다 낮은데도 불구하고 위너가 안된 경우
            if value[0] <= value[1]:
                #       유료배송인경우
                if value[2] > 0:
                    #         넷토피아가격 0.1% 내려도 노마진가격보다 높은경우 가격수정
                    if value[0] - (value[0] * 0.001) > value[3]:
                        prices.append(value[0] - (round((value[0] * 0.001), -1) + value[2]))  # 유료배송비 차감
                        text.append("")
                    else:
                        prices.append("")
                        text.append("확인필요")

                #       무료배송인경우
                else:
                    #         넷토피아가격 0.1% 내려도 노마진가격보다 높은경우 가격수정
                    if value[0] - (value[0] * 0.001) > value[3]:
                        prices.append(value[0] - round((value[0] * 0.001), -1))  # 배송비 차감하지 않음
                        text.append("")
                    else:
                        prices.append("")
                        text.append("확인필요")

            #    넷토피아 가격 > 쿠팡위너가격 보다 높은 경우
            else:
                #       유료배송인 경우
                if value[2] > 0:
                    #            쿠팡위너가격 1% 내려도 노마진가격보다 높은경우 가격수정
                    if value[1] - (value[0] * 0.01) > value[3]:
                        prices.append(value[1] - (round((value[0] * 0.01), -1) + value[2]))  # 유료배송비 차감
                        text.append("")
                    #           노마진이 9000원 미만인경우
                    elif value[3] < 9000:
                        #               노마진 - 700원이 < 쿠팡위너가격 낮은경우
                        if value[3] - 700 < value[1]:
                            prices.append(value[1] - (round((value[0] * 0.01), -1) + value[2]))  # 유료배송비 차감
                            text.append("최대수량적용")

                        else:
                            prices.append("")
                            text.append("확인필요")

                    else:
                        prices.append("")
                        text.append("확인필요")

                #       무료배송인 경우
                else:

                    if value[1] - (value[0] * 0.01) > value[3]:
                        prices.append(value[1] - round((value[0] * 0.01), -1))  # 배송비 차감하지 않음
                        text.append("")
                    else:
                        prices.append("")
                        text.append("확인필요")

        df_coupang["판매가수정"] = prices
        yield "data:75\n\n"

        margin = []

        for v in df_coupang[['판매가수정', '원가', '상품수량', '판매자 배송비(원)']].values:
            if v[3] > 0:
                if v[0] == "":
                    margin.append("")
                else:
                    margin.append((v[0] * 0.89) / (v[1] * v[2]))
            else:
                if v[0] == "":
                    margin.append("")
                else:
                    margin.append((v[0] * 0.89 - 1700) / (v[1] * v[2]))

        df_coupang["마진"] = margin

        df_coupang["비고2"] = text

        # def float2int(x):
        #     try:
        #         return str(round(int(x), 2))
        #     except:
        #         return ""

        for col in ["상품수량", "원가", "노마진"]:
            df_coupang[col] = df_coupang[col].apply(float3int)

        df_coupang.to_excel(os.path.join(output_dir, "price.xlsx"), index=False)
        yield "data:75\n\n"

        wb = openpyxl.load_workbook(os.path.join(output_dir, "price.xlsx"))
        ws = wb.active

        colors = ["상품수량", "원가", "노마진", '넷토피아가격', '쿠팡위너가격', '판매가수정', '비고']

        # colors = {"판매상태.1" : openpyxl.styles.colors.YELLOW, "잔여수량" : openpyxl.styles.colors.GREEN, "실제재고" : openpyxl.styles.colors.BLUE}

        for col in ws.columns:
            if col[0].value in colors:
                letter = col[0].column_letter
                for c in ws[letter]:
                    c.fill = PatternFill(fill_type='solid', start_color=YELLOW, end_color=YELLOW)

        for x in ws['T']:
            if x.value == "확인필요" or x.value == "최대수량적용":
                x.fill = PatternFill(fill_type='solid', start_color=YELLOW, end_color=YELLOW)

        for i in ws['R']:
            try:
                print(i.value)
                if float(i.value) >= float(1.1):
                    i.fill = PatternFill(fill_type='solid', start_color=GREEN, end_color=GREEN)
                elif float(i.value) >= float(1.03) and float(i.value) < float(1.1):
                    i.fill = PatternFill(fill_type='solid', start_color=YELLOW, end_color=YELLOW)
                else:
                    i.fill = PatternFill(fill_type='solid', start_color=RED, end_color=RED)
            except:
                pass

        wb.save(filename=os.path.join(output_dir, "price.xlsx"))
        yield "data:99\n\n"
        time.sleep(0.5)

        yield "data:100\n\n"
    return Response(makeWinner(), mimetype='text/event-stream')


if __name__ == "__main__":
    app.run()